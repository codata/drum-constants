"""
Produces a JSON representation of the CODATA constants Google spreadsheet: 
https://docs.google.com/spreadsheets/d/1m5Hm3uRsgDVXIarp7-AQqt2mYSvdk0Bvzgx3bvdMT6s/edit

The sheets holding the constant data values across versions are populated from the NIST published ASCII files.

"""
import argparse
import json
import openpyxl
import logging
import re
import requests

def download_gsheet(sheet_id, outfile):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    logging.info(f"Downloading from {url}")
    response = requests.get(url)
    if response.status_code == 200:
        with open(outfile, 'wb') as file:
            file.write(response.content)
            file.close()

def get_sheet_column_map(sheet, matches):
    "Return a map of column names to index positions"
    map = {}
    # header row
    for row in sheet.iter_rows(max_row=1):
        for index, cell in enumerate(row): 
            cell_value = str(cell.value)
            if (cell_value):
                for regex in matches:
                    if re.match(regex, cell_value, re.IGNORECASE):
                        # add to map
                        map_value = {'index':index, 'regex':regex}
                        map[cell_value] = map_value
    logging.debug(map)
    return map

def get_row_cell_value(row, map, column_name):
    index = map[column_name].get('index')
    if index is not None:
        return row[index].value

def get_sheet_entries(sheet, columns):
    """Parse a sheet and returns data for specified column names as a dictionary. An 'id' column must exist and is used as key."""
    entries = {}
    # get column map
    column_map = get_sheet_column_map(sheet, columns)
    # parse data
    for row in sheet.iter_rows(min_row=2):
        id = get_row_cell_value(row, column_map, "id")
        if id:
            data = {}
            for column in column_map:
                data[column] = get_row_cell_value(row, column_map, column)
            entries[id] = data
    return entries

def parse_workbook(filename):
    """Parses Excel file and populates the constants model"""
    logging.debug(f"workbook={filename}")
    wb = openpyxl.load_workbook(filename, data_only=True)

    output:dict = {
        "version":"0.1.0"
    }
    # QUANTITIES
    logging.info("Parsing quantities")
    quantities: list[dict] = []
    output['quantities'] = quantities
    sheet_quantities = get_sheet_entries(wb['Quantities'], ["id","name","notation","definition","same_as","has_parts","is_ratio","is_relationship"])
    quantities_index:dict={}
    for (id, entry) in sheet_quantities.items():
        # create
        quantity: dict = {"type":"Quantity"}
        quantity['id'] = f"{id}"
        quantity_ids = {}
        quantity['ids'] = quantity_ids
        if entry.get('name'):
            quantity['name'] = entry.get('name')
        if entry.get('definition'):
            quantity['definition'] = entry.get('definition')
        if entry.get('same_as'):
            sames = entry.get('same_as').split(',')
            quantity['same_as'] = [same.strip() for same in sames]
        if entry.get('has_parts'):
            sames = entry.get('has_parts').split(',')
            quantity['has_parts'] = [part.strip() for part in sames]
        if entry.get('is_ratio'):
            quantity['is_ratio'] = True
        if entry.get('is_relationship'):
            quantity['is_relationship'] = True
        quantity['constants'] = []
        # add
        quantities.append(quantity)
        quantities_index[id] = quantity

    # UNITS
    logging.info("Parsing units")
    units: list[dict] = []
    output['units'] = units
    sheet_units = get_sheet_entries(wb['Units'], ["id","name","unit_SI_expression","unit_SI_uri","unit_ucum","unit_uom"])
    units_index:dict={}
    for (id, entry) in sheet_units.items():
        # create
        unit: dict = {"type":"Unit"}
        unit['id'] = f"{id}"
        unit_ids = {}
        if entry.get('unit_SI_uri'):
            unit_ids['SI'] = entry.get('unit_SI_uri')
        if entry.get('unit_ucum'):
            unit_ids['UCUM'] = entry.get('unit_ucum')
        if entry.get('unit_uom'):
            unit_ids['UOM'] = entry.get('unit_uom')
        unit['ids'] = unit_ids
        # add
        units.append(unit)
        units_index[id] = unit

    # CONSTANTS 
    logging.info("Parsing constants")
    sheet_constants = get_sheet_entries(wb['Constants'], ["nist_id","id","name","name_bipm_en","name_bipm_fr","unit_nist","unit_id","quantity_id","qudt_id"])
    constants_index:dict = {}
    constants_quantities_map = {} # maps constants codata identifiers to quantities  to speed up version processing
    nist_constants_map = {} # maps nist identifiers to constants to speed up version processing
    for (id, entry) in sheet_constants.items():
        quantity_id = entry.get('quantity_id')
        quantity = quantities_index.get(quantity_id,{})
        if quantity:
            # create
            constant: dict = {"type":"Constant"}
            constant['id'] = f"{id}"
            constant_ids = {'NIST': entry.get('nist_id')}
            if entry.get('qudt_id'):
                constant_ids['QUDT'] = entry.get('qudt_id')
            constant['ids'] = constant_ids
            constant['name'] = entry.get('name')
            if entry.get('name_bipm_en'):
                constant['name_bipm_en'] = entry.get('name_bipm_en')
            if entry.get('name_bipm_fr'):
                constant['name_bipm_fr'] = entry.get('name_bipm_fr')
            if entry.get('unit_id'):
                constant['unit_id'] = entry['unit_id']
            constant['values'] = []
            # add
            quantity['constants'].append(constant)
            constants_index[id] = constant
            constants_quantities_map[id] = quantity['id']
            nist_constants_map[entry.get('nist_id')] = id
        else:
            logging.error(f"Quantity not found for Constant {id}")

    # VERSIONS/VALUES
    version_regex = r"v\d{4}" # match sheet name
    for name in wb.sheetnames:
        logging.debug(f"sheet_name={name}")
        # parse versions
        if re.match(version_regex, name, re.IGNORECASE):
            logging.info(f"Parsing version {name}")
            # process version
            sheet = wb[name]
            version_id = name[1:]
            data = get_sheet_entries(sheet, ["id","name","units","value_str","value_num","uncertainty_str","uncertainty_n","is_exact","is_truncated","exponent"])
            # add version to constants
            for (id, entry) in data.items():
                # lookup constant
                constant_id = nist_constants_map.get(id)
                constant = constants_index.get(constant_id,{})
                if not constant:
                    logging.error(f"Constant not found for {id}")
                    continue
                # lookup quantity
                quantity_id = constants_quantities_map.get(constant_id)
                quantity = quantities_index.get(quantity_id,{})
                if not quantity:
                    logging.error("Quantity not found for {id}")
                    continue
                # lookup/create versions property
                if not constant.get('values'):
                    constant['values'] = []
                constant_versions = constant['values']
                # add this version to the versions
                constant_version: dict = {"type":"ConstantVersion"}
                constant_versions.append(constant_version)
                # populate version data
                constant_value_ids = {}
                constant_version['ids'] = constant_value_ids
                constant_version['version'] = version_id
                constant_version['name'] = entry.get('name')
                constant_version['value'] = entry.get('value_str')
                uncertainty_str = entry.get('uncertainty_str')
                if uncertainty_str == '(exact)':
                    uncertainty_str = None
                constant_version['uncertainty'] = uncertainty_str
                if entry.get('exponent'):
                    constant_version['exponent'] = entry.get('exponent')
                if entry.get('units'):
                    constant_version['units'] = entry.get('units')
                constant_version['is_exact'] = entry.get('is_exact',False)
                constant_version['is_truncated'] = entry.get('is_truncated',False)
    return output

def main():
    sheet_filename = "codata_constants.xlsx"
    if not args.norefresh:
        download_gsheet("1m5Hm3uRsgDVXIarp7-AQqt2mYSvdk0Bvzgx3bvdMT6s", sheet_filename)
    constants = parse_workbook(sheet_filename)
    with open('codata_constants.json', 'w') as f:
        json.dump(constants, f, indent=4)

if __name__ == '__main__':
    global args
    parser = argparse.ArgumentParser()
    parser.add_argument("-nr","--norefresh", action="store_true", help="Skip downloading the spreadsheet and use existing file")
    parser.add_argument("-ll","--loglevel", help="Python logging level", default="INFO")
    args = parser.parse_args()
    print(args)

    logging.basicConfig(format='%(asctime)s %(levelname)s %(message)s')
    if args.loglevel:
        logging.getLogger().setLevel(args.loglevel.upper())

    logging.info(args)

    main()
