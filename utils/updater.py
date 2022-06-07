import argparse
from datetime import datetime
import json
import openpyxl
import logging
import numpy as np
import re
import requests

CODATA_CONSTANT_ID_PREFIX = "Constant"
CODATA_CONSTANT_INSTANCE_ID_PREFIX = "ConstantInstance"
CODATA_CONSTANT_VALUE_ID_PREFIX = "ConstantValue"

def download_gsheet(sheet_id, outfile):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    logging.info(f"Downloading from {url}")
    response = requests.get(url)
    if response.status_code == 200:
        with open(outfile, 'wb') as file:
            file.write(response.content)
            file.close()

def get_sheet_column_map(sheet, matches):
    "Return a map of column names to index positions"
    map = {}
    # header row
    for row in sheet.iter_rows(max_row=1):
        for index, cell in enumerate(row): 
            cell_value = str(cell.value)
            if (cell_value):
                for regex in matches:
                    if re.match(regex, cell_value, re.IGNORECASE):
                        # add to map
                        map_value = {'index':index, 'regex':regex}
                        map[cell_value] = map_value
    logging.debug(map)
    return map

def get_row_cell_value(row, map, column_name):
    index = map[column_name].get('index')
    if index is not None:
        return row[index].value

def get_sheet_entries(sheet, columns):
    """Parse a sheet and returns data for specified column names as a dictionary. An 'id' column must exist and is used as key."""
    entries = {}
    # get column map
    column_map = get_sheet_column_map(sheet, columns)
    # parse data
    for row in sheet.iter_rows(min_row=2):
        id = get_row_cell_value(row, column_map, "id")
        if id:
            data = {}
            for column in column_map:
                data[column] = get_row_cell_value(row, column_map, column)
            entries[id] = data
    return entries

def parse_workbook(filename):
    """Parses Excel file and populates the constants model"""
    logging.debug(f"workbook={filename}")
    wb = openpyxl.load_workbook(filename, data_only=True)

    output = {
        "version":"0.1.0"
    }
    constants = []
    output['constants'] = constants
    # CONSTANTS DEFINITIONS
    logging.info("Parsing constants")
    sheet_constants = get_sheet_entries(wb['ConstantsDefinitions'], ["id","name","definition","dimensionless"])
    constants_index={}
    for (id, entry) in sheet_constants.items():
        # create
        constant = {"type":"ConstantDefinition"}
        constant['id'] = f"{id}"
        codata_id = f"{CODATA_CONSTANT_ID_PREFIX}:{id}"
        constant_ids = {'codata':codata_id}
        constant['ids'] = constant_ids
        if entry.get('name'):
            constant['name'] = entry.get('name')
        if entry.get('definition'):
            constant['definition'] = entry.get('definition')
        if entry.get('dimensionless'):
            constant['is_dimensionless'] = True
        constant['instances'] = []
        # add
        constants.append(constant)
        constants_index[codata_id] = constant
    # CONSTANTS INSTANCES
    logging.info("Parsing constants instances")
    sheet_constants_instances = get_sheet_entries(wb['ConstantsInstances'], ["id","name","units_si","units_ucum","units_uom","constant_id","qudt_id"])
    constants_instances_index = {}
    constants_instances_map = {} # maps constant units codata identifiers to their constant to speed up version processing
    for (id, entry) in sheet_constants_instances.items():
        constant_codata_id =  f"{CODATA_CONSTANT_ID_PREFIX}:{entry['constant_id']}"
        constant = constants_index.get(constant_codata_id)
        if constant:
            # create
            constant_instance = {"type":"ConstantInstance"}
            constant_instance['id'] = f"{id}"
            codata_id = f"{CODATA_CONSTANT_INSTANCE_ID_PREFIX}:{id}"
            constant_instance_ids = {'CODATA': codata_id}
            if entry.get('qudt_id'):
                constant_instance_ids['QUDT'] = entry.get('qudt_id')
            constant_instance['ids'] = constant_instance_ids
            constant_instance['name'] = entry.get('name')
            constant_instance['units'] = {}
            if entry.get('units_si'):
                constant_instance['units']['SI'] = entry['units_si']
            if entry.get('units_ucum'):
                constant_instance['units']['UCUM'] = entry['units_ucum']
            if entry.get('units_uom'):
                constant_instance['units']['UOM'] = entry['units_uom']
            constant_instance['versions'] = []
            # add
            constant['instances'].append(constant_instance)
            constants_instances_index[codata_id] = constant_instance
            constants_instances_map[codata_id] = constant_codata_id
        else:
            logging.error(f"Constant not found for ConstantInstance {id}")
    # VERSIONS
    version_regex = "v\d{4}" # match sheet name
    for name in wb.sheetnames:
        logging.debug(f"sheet_name={name}")
        # parse versions
        if re.match(version_regex, name, re.IGNORECASE):
            logging.info(f"Parsing version {name}")
            # process version
            sheet = wb[name]
            version_id = name[1:]
            data = get_sheet_entries(sheet, ["id","name","units_si","value_str","value_num","uncertainty_str","uncertainty_n","ellipsis","exponent"])
            # add version to constants
            for (id, entry) in data.items():
                # find constant unit and constant
                constant_instance_codata_id = f"{CODATA_CONSTANT_INSTANCE_ID_PREFIX}:{id}"
                constant_codata_id = constants_instances_map.get(constant_instance_codata_id)
                if not constant_codata_id:
                    logging.error(f"Constant identifier not found for {id}")
                    continue
                # lookup constant
                constant = constants_index.get(constant_codata_id)
                if not constant:
                    logging.error("Constant not found for {id}")
                    continue
                # lookup constant units
                constant_instance = constants_instances_index.get(constant_instance_codata_id)
                if not constant_instance:
                    logging.error(f"ConstantUnits not found for {id}")
                    continue
                # lookup/create versions property
                if not constant_instance.get('versions'):
                    constant_instance['versions'] = []
                constant_instance_versions = constant_instance['versions']
                # add this version to the versions
                codata_id = constant_instance_codata_id.replace(CODATA_CONSTANT_INSTANCE_ID_PREFIX,CODATA_CONSTANT_VALUE_ID_PREFIX)+":"+version_id
                constant_instance_version = {"type":"ConstantVersion"}
                constant_instance_versions.append(constant_instance_version)
                # populate version data
                constant_value_ids = {'codata':codata_id}
                constant_instance_version['ids'] = constant_value_ids
                constant_instance_version['version'] = version_id
                constant_instance_version['name'] = entry.get('name')
                if entry.get('units'):
                    constant_instance_version['units'] = entry.get('units')
                constant_instance_version['value'] = entry.get('value_str')
                uncertainty_str = entry.get('uncertainty_str')
                uncertainty = float(uncertainty_str) if uncertainty_str and uncertainty_str != '(exact)' else None
                #print(type(entry.get('uncertainty_n')), uncertainty_str, uncertainty)
                constant_instance_version['uncertainty'] = "{:.1e}".format(uncertainty) if uncertainty else None
                if entry.get('exponent'):
                    constant_instance_version['exponent'] = entry.get('exponent')
                if entry.get('ellipsis'):
                    constant_instance_version['ellipsis'] = True
    return output

def main():
    sheet_filename = "codata_constants.xlsx"
    download_gsheet("1m5Hm3uRsgDVXIarp7-AQqt2mYSvdk0Bvzgx3bvdMT6s", sheet_filename)
    constants = parse_workbook(sheet_filename)
    with open('codata_constants.json', 'w') as f:
        json.dump(constants, f, indent=4)

if __name__ == '__main__':
    global args
    parser = argparse.ArgumentParser()
    parser.add_argument("-ll","--loglevel", help="Python logging level", default="INFO")
    args = parser.parse_args()
    print(args)

    logging.basicConfig(format='%(asctime)s %(levelname)s %(message)s')
    if args.loglevel:
        logging.getLogger().setLevel(args.loglevel.upper())

    logging.info(args)

    main()
